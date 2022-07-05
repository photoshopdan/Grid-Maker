from genericpath import getsize
import os
import sys
import glob
import numpy as np
from PIL import Image, ImageFont, ImageDraw, ImageColor, ImageCms
from iptcinfo3 import IPTCInfo
import csv
from openpyxl import load_workbook
from PyQt5.QtCore import (QCoreApplication, Qt, QObject, pyqtSignal, pyqtSlot,
                          QRunnable, QThreadPool)
from PyQt5.QtWidgets import (QWidget, QApplication, QProgressBar, QMainWindow,
                             QPushButton, QComboBox, QLabel, QStyleFactory,
                             QColorDialog, QMessageBox, QAction, QLineEdit,
                             QCheckBox, QProgressDialog, QListWidget,
                             QAbstractItemView)
from PyQt5.QtGui import (QIcon, QColor)


class Layout():

    def __init__(self, size, instructions, aspect):
        self.size_dict = {
            '18" x 12"': (5400, 3600),
            '12" x 8"': (3600, 2400),
            '15" x 12"': (4500, 3600),
            '10" x 8"': (3000, 2400),
            '30" x 20"': (9000, 6000)
            }
        self.ratio_dict = {
            1.5: '32',
            1.33: '43'
            }
        self.canvas_size = self.size_dict[size]
        self.canvas_graphics = self.csvread(os.path.join(ROOT_DIR,
                                            'Layouts\\'+instructions+'_'+
                                            str(self.canvas_size[0])+'x'+
                                            str(self.canvas_size[1])+
                                            '_canvas_graphics'+'_'+
                                            self.ratio_dict[aspect]+'.csv'))
        self.aspect, self.instructions = aspect, instructions

        if instructions != 'Full-page grid layout':
            self.imsize = self.csvread(os.path.join(ROOT_DIR,
                                        'Layouts\\'+instructions+'_'+
                                       str(self.canvas_size[0])+'x'+
                                       str(self.canvas_size[1])+
                                       '_imsize'+'_'+self.ratio_dict[aspect]+
                                       '.csv'))
            
            self.implace = self.csvread(os.path.join(ROOT_DIR,
                                        'Layouts\\'+instructions+'_'+
                                        str(self.canvas_size[0])+'x'+
                                        str(self.canvas_size[1])+
                                        '_implace'+'_'+self.ratio_dict[aspect]+
                                        '.csv'))

    def make_canvas(self):
        canv = Image.new('RGB', self.canvas_size, '#ffffff')
        return canv

    def make_graphics(self, canv, colour1, colour2, colour3, n):
        draw = ImageDraw.Draw(canv)
        if self.instructions == 'Full-page grid layout':
            for p, c in list(zip(eval(self.canvas_graphics[0][0]),
                                 eval(self.canvas_graphics[0][1]))):
                draw.rectangle(p, fill = c, outline = None)
        else:
            for p, c in list(zip(eval(self.canvas_graphics[n-1][0]),
                                 eval(self.canvas_graphics[n-1][1]))):
                draw.rectangle(p, fill = c, outline = None)

    def calc_size(self, n, im_w_pad, im_h_pad, custom, c_w, c_h):
        im_w, im_h = 200, 200 * self.aspect
        if custom == False:
            if self.instructions == 'Full-page grid layout':
                c_tlx, c_tly, c_brx, c_bry = eval(self.canvas_graphics[0][7])
            else:
                c_tlx, c_tly, c_brx, c_bry = eval(self.canvas_graphics[n-1][7])
            c_w, c_h = c_brx - c_tlx, c_bry - c_tly
        else:
            c_tlx, c_tly = 0, 0
        ratio_padded = ((im_w + (im_w * im_w_pad))
                        / (im_h + (im_w * im_h_pad)))
        frame_im_ratio = (((im_w + (im_w * im_w_pad)) * c_h)
                          / ((im_h + (im_h_pad * im_w)) * c_w))
        columns = int(round(np.sqrt(n / frame_im_ratio)))
        rows = int(np.ceil(n / columns))
        if columns - (columns * rows - n) < columns / 3:
            columns = columns + 1
            rows = rows - 1
            if columns == 0:
                columns = 1
            if rows == 0:
                rows = 1
        if n == 1:
            columns, rows = 1, 1
        if n == 2:
            columns, rows = 2, 1
        if n == 3:
            columns, rows = 3, 1
        if n == 4:
            columns, rows = 4, 1
        if n == 5:
            columns, rows = 3, 2
        max_w, max_h = c_w / columns, c_h / rows
        resize_w_pad = min(max_w, max_h * ratio_padded)
        resize_h_pad = min(max_w / ratio_padded, max_h)

        resize_w = int(round(resize_w_pad
                             * (im_w / (im_w + (im_w * im_w_pad)))))
        resize_h = int(round(resize_h_pad
                             * (im_h / (im_h + (im_w * im_h_pad)))))

        return (resize_w, resize_h, columns, rows, max_w, max_h, c_tlx, c_tly)


    def image_place(self, canv, n, i, im, grid_data):
        if self.instructions == 'Full-page grid layout':
            im_w, im_h, columns, rows, max_w, max_h, c_tlx, c_tly = grid_data
            if np.ceil((i + 1) / columns) == rows:
                right_shift = (((columns / 2)
                                - ((columns - (columns * rows - n)) / 2))
                               * max_w)
                pos_x = round(((max_w * ((i % columns) + 1)
                                - im_w - ((max_w - im_w) / 2)) + c_tlx)
                              + right_shift)
                pos_y = round((max_h * np.ceil((i + 1) / columns)
                               - max_h + ((max_h - im_h) / 2)) + c_tly)
                canv.paste(im, (int(pos_x), int(pos_y)))
            else:
                pos_x = round((max_w * ((i % columns) + 1)
                               - im_w - ((max_w - im_w) / 2)) + c_tlx)
                pos_y = round((max_h * np.ceil((i + 1) / columns)
                               - max_h + ((max_h - im_h) / 2)) + c_tly)
                canv.paste(im, (int(pos_x), int(pos_y)))

        else:
            canv.paste(im, eval(self.implace[n-1][i]))

    def image_resize(self, im, n, calc_size):
        if self.instructions == 'Full-page grid layout':
            img = im.resize(calc_size,
                            resample = Image.HAMMING,
                            reducing_gap = 2.0)
        else:
            img = im.resize(eval(self.imsize[0][n-1]),
                            resample = Image.HAMMING,
                            reducing_gap = 2.0)
            
        return img

    def image_resize_caption(self, im, n, ratio, custom, calc_size, txt, col):
        if self.instructions == 'Full-page grid layout':
            canvas_x = Image.new('RGB',
                                 calc_size,
                                 color = '#ffffff')
            if custom == True:
                caption_resize_x = int((calc_size[1]
                                        - (txt * 2.8)) / ratio)
                caption_resize_y = int(calc_size[1]
                                       - (txt * 2.8))
            else:
                caption_resize_x = int((calc_size[1]
                                        - (eval(self.canvas_graphics[0][8])
                                           * 2.8)) / ratio)
                caption_resize_y = int(calc_size[1]
                                       - (eval(self.canvas_graphics[0][8])
                                           * 2.8))
            img = im.resize((caption_resize_x, caption_resize_y),
                            resample = Image.HAMMING,
                            reducing_gap = 2.0)
            img = self.keyline(img, col)
            canvas_x.paste(img, (int((calc_size[0] / 2)
                                     - caption_resize_x / 2), 0))
        else:
            canvas_x = Image.new('RGB',
                                 (eval(self.imsize[0][n-1])),
                                 color = '#ffffff')
            caption_resize_x = int(((eval(self.imsize[0][n-1])[1])
                                    * 0.76) / ratio)
            caption_resize_y = int((eval(self.imsize[0][n-1])[1])
                                   * 0.76)
            img = im.resize((caption_resize_x, caption_resize_y),
                            resample = Image.HAMMING,
                            reducing_gap = 2.0)
            img = self.keyline(img, col)
            canvas_x.paste(img, (int((canvas_x.width / 2)
                                     - caption_resize_x / 2), 0))
            
        return canvas_x

    def keyline(self, im, col):
        if col == '#ffffff':
            return im
        else:
            img_w, img_h = im.size
            img_array = np.array(im)
            colour = ImageColor.getrgb(col)

            img_array[0:4] = colour
            img_array[:, 0:4] = colour
            img_array[:, (img_w - 4):img_w] = colour
            img_array[(img_h - 4):img_h] = colour

            return Image.fromarray(img_array)

    def logo_place(self, canv, im, n):
        if im.mode == 'RGBA':
            if self.instructions == 'Full-page grid layout':
                canv.paste(im, eval(self.canvas_graphics[0][3]), mask = im)
            else:
                canv.paste(im, eval(self.canvas_graphics[n-1][3]), mask = im)
        else:
            if self.instructions == 'Full-page grid layout':
                canv.paste(im, eval(self.canvas_graphics[0][3]))
            else:
                canv.paste(im, eval(self.canvas_graphics[n-1][3]))

    def text_place(self, canv, schoolname, classname, logo_width, n, col4):
        draw = ImageDraw.Draw(canv)
        scale_factor_dict = {(5400, 3600): 1.08,
                             (3600, 2400): 1.08,
                             (4500, 3600): 1.18,
                             (3000, 2400): 1.18,
                             (9000, 6000): 1.08}
        scale_factor = scale_factor_dict[self.canvas_size]
        if self.instructions == 'Full-page grid layout':
            font = ImageFont.truetype("GOTHIC.TTF",
                                      eval(self.canvas_graphics[0][4]))
            class_text_width = font.getsize(classname)
            school_text_width = font.getsize(schoolname)
            full_width = (class_text_width[0]
                          + school_text_width[0]
                          + logo_width)
            canv_space = (eval(self.canvas_graphics[0][6])[0]
                          + class_text_width[0]
                          - eval(self.canvas_graphics[0][3])[0])
            school_text_position = eval(self.canvas_graphics[0][5])
            class_text_position = eval(self.canvas_graphics[0][6])
            if (full_width * scale_factor) > canv_space:
                decrease_multiplier = canv_space / (full_width * scale_factor)
                font = ImageFont.truetype("GOTHIC.TTF",
                                          int(eval(self.canvas_graphics[0][4])*
                                              decrease_multiplier))
                school_text_width_copy = school_text_width
                class_text_width_copy = class_text_width
                school_text_width = font.getsize(schoolname)
                class_text_width = font.getsize(classname)
                school_text_position = list(eval(self.canvas_graphics[0][5]))
                class_text_position = list(eval(self.canvas_graphics[0][6]))
                school_text_position[1] += (school_text_width_copy[1]
                                            - school_text_width[1]) / 2
                class_text_position[1] += (class_text_width_copy[1]
                                           - class_text_width[1]) / 2
                
            draw.text(tuple(school_text_position),
                      schoolname, font = font, fill = col4)
            draw.text(tuple(class_text_position),
                      classname, font = font, fill = col4)
        else:
            font = ImageFont.truetype("GOTHIC.TTF",
                                      eval(self.canvas_graphics[n-1][4]))
            class_text_width = font.getsize(classname)
            school_text_width = font.getsize(schoolname)
            full_width = (class_text_width[0]
                          + school_text_width[0]
                          + logo_width)
            canv_space = (eval(self.canvas_graphics[n-1][6])[0]
                          + class_text_width[0]
                          - eval(self.canvas_graphics[n-1][3])[0])
            school_text_position = eval(self.canvas_graphics[n-1][5])
            class_text_position = eval(self.canvas_graphics[n-1][6])
            if (full_width * scale_factor) > canv_space:
                decrease_multiplier = canv_space / (full_width * scale_factor)
                font = ImageFont.truetype("GOTHIC.TTF",
                                          int(eval(self.canvas_graphics[n-1]
                                                   [4])*
                                              decrease_multiplier))
                school_text_width_copy = school_text_width
                class_text_width_copy = class_text_width
                school_text_width = font.getsize(schoolname)
                class_text_width = font.getsize(classname)
                school_text_position = list(eval(self.canvas_graphics[n-1][5]))
                class_text_position = list(eval(self.canvas_graphics[n-1][6]))
                school_text_position[1] += (school_text_width_copy[1]
                                            - school_text_width[1]) / 2
                class_text_position[1] += (class_text_width_copy[1]
                                           - class_text_width[1]) / 2
            
            draw.text(tuple(school_text_position),
                      schoolname, font = font, fill = col4)
            draw.text(tuple(class_text_position),
                      classname, font = font, fill = col4)

    def name_text_place(self, canv, fname, sname, n, custom, txt, col5):
        draw = ImageDraw.Draw(canv)
        if self.instructions == 'Full-page grid layout':
            if custom == True:
                fname_font = ImageFont.truetype("GOTHIC.TTF", txt)
                sname_font = ImageFont.truetype("GOTHIC.TTF", txt)
            else:
                fname_font = ImageFont.truetype(
                    "GOTHIC.TTF", eval(self.canvas_graphics[0][8]))
                sname_font = ImageFont.truetype(
                    "GOTHIC.TTF", eval(self.canvas_graphics[0][8]))
        else:
            fname_font = ImageFont.truetype(
                "GOTHIC.TTF", eval(self.canvas_graphics[n-1][7]))
            sname_font = ImageFont.truetype(
                "GOTHIC.TTF", eval(self.canvas_graphics[n-1][7]))
        fname_text_width, _ = fname_font.getsize(fname)
        sname_text_width, _ = sname_font.getsize(sname)
        canv_w, canv_h = canv.size
        if self.instructions == 'Full-page grid layout':
            if custom == True:
                fname_position = (canv_h - (txt * 2.4)) / canv_h
                sname_position = (canv_h - (txt * 1.2)) / canv_h
            else:
                fname_position = ((canv_h - (eval(self.canvas_graphics[0][8])
                                             * 2.4))
                                  / canv_h)
                sname_position = ((canv_h - (eval(self.canvas_graphics[0][8])
                                             * 1.2))
                                  / canv_h)
        else:
            fname_position = 0.775
            sname_position = 0.877
        if fname_text_width > canv_w:
            decrease_multiplier = canv_w / fname_text_width
            if self.instructions == 'Full-page grid layout':
                if custom == True:
                    fname_font = ImageFont.truetype(
                        "GOTHIC.TTF", int(txt * decrease_multiplier))
                else:
                    fname_font = ImageFont.truetype(
                        "GOTHIC.TTF", int(eval(self.canvas_graphics[0][8])
                                          * decrease_multiplier))
            else:
                fname_font = ImageFont.truetype(
                    "GOTHIC.TTF", int(eval(self.canvas_graphics[n-1][7])
                                      * decrease_multiplier))
            fname_text_width, _ = fname_font.getsize(fname)
            fname_position = fname_position + ((1 - decrease_multiplier) / 20)
        if sname_text_width > canv_w:
            decrease_multiplier = canv_w / sname_text_width
            if self.instructions == 'Full-page grid layout':
                if custom == True:
                    sname_font = ImageFont.truetype(
                        "GOTHIC.TTF", int(txt * decrease_multiplier))
                else:
                    sname_font = ImageFont.truetype(
                        "GOTHIC.TTF", int(eval(self.canvas_graphics[0][8])
                                          * decrease_multiplier))
            else:
                sname_font = ImageFont.truetype(
                    "GOTHIC.TTF", int(eval(self.canvas_graphics[n-1][7])
                                      * decrease_multiplier))
            sname_text_width, _ = sname_font.getsize(sname)
            sname_position = sname_position + ((1 - decrease_multiplier) / 20)
            
        draw.text((int((canv_w / 2) - (fname_text_width / 2)),
                   int(canv_h * fname_position)),
                  fname, font = fname_font, fill = col5)
        draw.text((int((canv_w / 2) - (sname_text_width / 2)),
                   int(canv_h * sname_position)),
                  sname, font = sname_font, fill = col5)

    def make_gv(self, img):
        if any([self.canvas_size == (5400, 3600),
                self.canvas_size == (3600, 2400)]):
            img_gv = img.resize((1800, 1200),
                                resample = Image.HAMMING,
                                reducing_gap = 2.0)
            with Image.open(os.path.join(
                ROOT_DIR, 'GV Templates', '18x12.jpg')) as gv:
                gv_canvas = gv
                gv_canvas.load()
            gv_canvas.paste(img_gv, (0, 300))
            gv_canvas = self.watermark(gv_canvas)

            return gv_canvas
        elif self.canvas_size == (9000, 6000):
            img_gv = img.resize((3000, 2000),
                                resample = Image.HAMMING,
                                reducing_gap = 2.0)
            with Image.open(os.path.join(
                ROOT_DIR, 'GV Templates', '30x20.jpg')) as gv:
                gv_canvas = gv
                gv_canvas.load()
            gv_canvas.paste(img_gv, (0, 500))
            gv_canvas = self.watermark(gv_canvas)

            return gv_canvas
        else:
            img_gv = img.resize((1500, 1200),
                                resample = Image.HAMMING,
                                reducing_gap = 2.0)
            with Image.open(os.path.join(
                ROOT_DIR, 'GV Templates', '15x12.jpg')) as gv:
                gv_canvas = gv
                gv_canvas.load()
            gv_canvas.paste(img_gv, (0, 300))
            gv_canvas = self.watermark(gv_canvas)

            return gv_canvas

    def watermark(self, img):
        img_size = img.size
        txt_size = int(img_size[0] * 0.0806)
        img_watermark = img.convert('RGBA')
        txt_canv = Image.new('RGBA', img_size,
                             (255, 255, 255, 0))
        draw = ImageDraw.Draw(txt_canv)
        font = ImageFont.truetype('GOTHIC.TTF', txt_size)
        draw.text((int(img_size[0] * 0.04),
                   int((img_size[1] / 3) - txt_size * 0.75)),
                  'www.pret-a-portrait.net',
                  font = font, fill = (180, 180, 180, 30))
        img_watermark = Image.alpha_composite(img_watermark, txt_canv)
        img_watermark = img_watermark.convert('RGB')

        return img_watermark

    def csvread(self, path):
        with open(path) as f:
            f_csv = csv.reader(f)
            headers = next(f_csv)
            csvvar = []
            for row in f_csv:
                csvvar.append(row)
        return csvvar


class WorkerSignals(QObject):
    loading = pyqtSignal(int)
    progress = pyqtSignal(int)
    error01 = pyqtSignal(bool)
    error02 = pyqtSignal(bool)
    error03 = pyqtSignal(bool)
    error04 = pyqtSignal(bool)
    error05 = pyqtSignal(bool)
    error06 = pyqtSignal(int, str)
    error07 = pyqtSignal(bool)
    error08 = pyqtSignal(int, str)
    go = pyqtSignal(bool)


class JobRunner(QRunnable):
    
    signals = WorkerSignals()

    def __init__(self):
        super().__init__()

        self.logostatus = False
        self.customstatus = False
        self.inputtextsize = []
        self.inputsize = []
        self.inputlayout = []
        self.inputorder = []
        self.inputname = []
        self.inputcolour1 = []
        self.inputcolour2 = []
        self.inputcolour3 = []
        self.inputcolour4 = []
        self.inputcolour5 = []
        self.inputcolour6 = []
        self.gvstatus = False
        self.inputschool = []
        self.inputdate = []
        self.inputwidth = 0
        self.inputheight = 0
        self.inputwidthpad = 0.2
        self.inputheightpad = 0.2

    @pyqtSlot()
    def run(self):
        # Produce list of paths of all the required files.
        # Prompt if they can't be found.
        self.signals.loading.emit(0)
        self.folder = [r'C:\Users\Dan\Documents\Python\Grid Maker\Benchmark Set\3-2 Ratio']#sys.argv[1:]
        self.img_list = []
        self.img_id_list = []
        if len(self.folder) != 1:
            self.signals.error01.emit(True)
            raise UserWarning('Exception passed to user.')
        if not os.path.isdir(self.folder[0]):
            self.signals.error01.emit(True)
            raise UserWarning('Exception passed to user.')
        folder_short = (self.folder[0][self.folder[0].rindex('\\'):])[1:]
        try:
            spreadsheet = glob.glob('%s\\**\\*.xlsx' % self.folder[0],
                                    recursive = True)[0]
            images = glob.glob('%s\\**\\*.jpg' % self.folder[0],
                               recursive = True) +\
                     glob.glob('%s\\**\\*.png' % self.folder[0],
                               recursive = True)
        except IndexError:
            self.signals.error02.emit(True)
            raise UserWarning('Exception passed to user.')
        if not images:
            self.signals.error02.emit(True)
            raise UserWarning('Exception passed to user.')
            self.signals.loading.emit(20)
        else:
            # Create list of images in folder and segregate logo.
            print('Loading images')
            for img_name in images:
                img_name_short = ((img_name[img_name.rindex('\\'):])[1:])
                if 'logo.' in img_name_short.casefold():
                    self.logo = img_name
                    self.logostatus = True
                else:
                    self.img_list.append(img_name)
                    self.img_id_list.append(img_name_short[
                        :img_name_short.rindex('.')])
                try:
                    check_for_image = self.img_list[0]
                except IndexError:
                    self.signals.error02.emit(True)
                    raise UserWarning('Exception passed to user.')
            self.signals.loading.emit(40)

            # Read spreadsheet.
            print('Processing spreadsheet\n')
            spreadsheet_data = (self.excel_to_list(spreadsheet,
                                                   ('H',
                                                    'AJ',
                                                    'AL',
                                                    'AQ',
                                                    'AR',
                                                    'AT'))[1:])

            # Put teachers before students by renaming sitting type code.
            try:
                for x in spreadsheet_data:
                    if x[0] in ['6', '73', '79', '93', '172', '173', '221',
                                '240', '259', '278', '297', '316', '351',
                                '359', '391', '393', '432', '512', '524',
                                '534', '549', '554', '558', '600', '621']:
                        x[0] = '1'
                    else:
                        x[0] = '2'
            except KeyError:
                self.signals.error07.emit(True)
                raise UserWarning('Exception passed to user.')

            # Check for missing data.
            spreadsheet_id_list = [x[1] for x in spreadsheet_data]
            start_at = -1
            match_indices = []
            for x in spreadsheet_id_list:
                if x in self.img_id_list:
                    loc = spreadsheet_id_list.index(x, start_at + 1)
                    match_indices.append(loc)
                    start_at = loc
                else:
                    pass
            
            self.spreadsheet_data = [spreadsheet_data[x] for
                                     x in match_indices]
            spreadsheet_id_list = [spreadsheet_id_list[x] for
                                   x in match_indices]

            csv_data = [[x + '.jpg'] for x in self.img_id_list if
                        x not in spreadsheet_id_list]
            self.signals.loading.emit(60)
            # Prompt user to update spreadsheet if data is missing.
            if csv_data:
                csv_path = self.folder[0] + '\\Missing data.csv'
                headers = ['Images without corresponding spreadsheet data']
                self.csvwrite(csv_path, headers, 'w')
                for x in csv_data:
                    self.csvwrite(csv_path, x, 'a')
                self.signals.error03.emit(True)
                raise UserWarning('Exception passed to user.')
                self.signals.loading.emit(80)
            else:
                # Prompt user if images are of different ratios
                # or if the ratios are not 2:3 or 4:3.
                self.img_ratio_list = []
                for ims in self.img_list:
                    try:
                        with Image.open(ims) as im:
                            img_w, img_h = im.size
                            self.img_ratio_list.append(round(img_h/img_w, 2))
                    except(IOError, SyntaxError):
                        pass

                if self.img_ratio_list[1:] != self.img_ratio_list[:-1]:
                    self.signals.error04.emit(True)
                    raise UserWarning('Exception passed to user.')
                elif (self.img_ratio_list[0] != 1.5 and
                      self.img_ratio_list[0] != 1.33):
                    self.signals.error04.emit(True)
                    raise UserWarning('Exception passed to user.')
                else:
                    # Check if classes have more than one person with the same
                    # name, suggesting duplicates with independant IDs.
                    dupecheck = [[x[2], x[3] + ' ' + x[4]] for x
                                 in self.spreadsheet_data]
                    seen = []
                    dupes = set()
                    for x in dupecheck:
                        if x in seen:
                            dupes.add(x[1])
                        else:
                            seen.append(x)
                    if dupes:
                        self.signals.error08.emit(len(dupes), '\n'.join(dupes))
                        raise UserWarning('Exception passed to user.')
                    self.signals.loading.emit(100)
                    
                    # Feed class list to menu. Start menu.
                    self.class_set = set([x[2] for x in self.spreadsheet_data])
                    self.signals.go.emit(True)


    def commence(self):
        
        overflow_warning = []
        
        # Sort spreadsheet data according to user-selected attribute.
        if self.inputorder == 'Surname':
            sort1 = 0
            sort2 = 4
            sort3 = 3
        if self.inputorder == 'First name':
            sort1 = 0
            sort2 = 3
            sort3 = 4
        if self.inputorder == 'Student ID':
            sort1 = 0
            sort2 = 1
            sort3 = 4
        if self.inputorder == 'Custom (column AT)':
            sort1 = 5
            sort2 = 4
            sort3 = 3

        spreadsheet_data = sorted(self.spreadsheet_data,
                                  key = lambda x: (x[2].casefold(),
                                                   x[sort1].casefold(),
                                                   x[sort2].casefold(),
                                                   x[sort3].casefold()))

        # Make images closer when name captions used.
        if self.inputname == True:
            if self.customstatus == False:
                self.inputwidthpad = 0.1
                self.inputheightpad = 0.1
        # Convert points to pixels.
        if not self.inputtextsize:
            self.inputtextsize = 42
        else:
            self.inputtextsize = int(round(float(self.inputtextsize)
                                           * (25 / 6)))

        # Split into a list for each grade.
        school = []
        for i in self.class_set:
            school.append([x for x in spreadsheet_data if x[2] == i])

        # Load layout data
        layout = Layout(self.inputsize,
                        self.inputlayout,
                        self.img_ratio_list[0])
        profile = ImageCms.ImageCmsProfile(os.path.join(ROOT_DIR, 'ICC Profile/'
                                           + 'sRGB Color Space Profile.icm'))
        img_profile = profile.tobytes()
        bad_characters = ['\\', '/', ':', '*', '?', '"', '<', '>', '|', '\n']
        # Create folder in which grids will be saved.
        gridfolder = os.path.join(self.folder[0], 'Grids')
        if not os.path.exists(gridfolder):
            os.mkdir(gridfolder)
        # Create a folder for GoVisually versions if required.
        if self.gvstatus == True:
            gvfolder = os.path.join(self.folder[0], 'Grids', 'GoVisually')
            if not os.path.exists(gvfolder):
                os.mkdir(gvfolder)
        # Prepare logo.
        try:
            with Image.open(self.logo) as im:
                self.logo = im
                self.logo.load()
            logo_w, logo_h = self.logo.size
            logo_height = eval(layout.canvas_graphics[0][2])
            logo_width = int(logo_height/(logo_h/logo_w))
            self.logo = self.logo.resize((logo_width, logo_height))
        except AttributeError:
            self.logo = []
            logo_width = -90
            
        # Construct canvas.
        if self.customstatus == True:
            canvas = Image.new('RGB', (int(self.inputwidth),
                                       int(self.inputheight)),
                               '#ffffff')
        else:
            canvas = layout.make_canvas()
        prog_n = 0
        for grade in school:
            if self.inputlayout != 'Full-page grid layout':
                if len(grade) > 40:
                    overflow_warning.append(grade[0][2])
                    continue
            imgs = []
            firstnames = []
            surnames = []
            for student in grade:
                imgs.append([self.img_id_list.index(x) for x in
                             self.img_id_list if x == student[1]])
                if self.inputname == True:
                    if student[3] == 'None':
                        firstnames.append(' ')
                    else:
                        if student[3].isupper():
                            firstnames.append(student[3].title())
                        else:
                            firstnames.append(student[3])
                    if student[4] == 'None':
                        surnames.append(' ')
                    else:
                        if student[4].isupper():
                            surnames.append(student[4].title())
                        else:
                            surnames.append(student[4])
            files = []
            for i in imgs:
                files.append([self.img_list[x] for x in i])
            n = len(files)
            if self.inputlayout == 'Full-page grid layout':
                calc_size_output = layout.calc_size(n,
                                                    float(self.inputwidthpad),
                                                    float(self.inputheightpad),
                                                    int(self.customstatus),
                                                    int(self.inputwidth),
                                                    int(self.inputheight))
                img_calc_size = (calc_size_output[0], calc_size_output[1])
            canvas_i = canvas.copy()
            if self.customstatus == False:
                layout.make_graphics(canvas_i,
                                     self.inputcolour1,
                                     self.inputcolour2,
                                     self.inputcolour3,
                                     n)
            
            # Add images.
            for file in files:
                file_i = files.index(file)
                try:
                    with Image.open(file[0]) as im:
                        img = im
                        img.load()
                # Replace corrupt images with green frame.
                except(IOError, SyntaxError):
                    img = Image.new('RGB', (200, 300), '#00ff00')
                img_w, img_h = img.size
                img_ratio = round(img_h/img_w, 2)
                if img_ratio != 1.5 and img_ratio != 1.33:
                    print('IMAGE IS WRONG ASPECT RATIO!')
                    sys.exit()
                if self.inputname == False:
                    if self.inputlayout == 'Full-page grid layout':
                        img = layout.image_resize(img, n, img_calc_size)
                        img = layout.keyline(img, self.inputcolour6)
                        layout.image_place(canvas_i, n, file_i, img,
                                           calc_size_output)
                    else:
                        img = layout.image_resize(img, n, None)
                        img = layout.keyline(img, self.inputcolour6)
                        layout.image_place(canvas_i, n, file_i, img, None)
                    prog_n += (100/n)/len(school)
                    self.signals.progress.emit(int(prog_n))
                else:
                    if self.inputlayout == 'Full-page grid layout':
                        img = layout.image_resize_caption(img, n,
                                                          img_ratio,
                                                          self.customstatus,
                                                          img_calc_size,
                                                          self.inputtextsize,
                                                          self.inputcolour6)
                    else:
                        img = layout.image_resize_caption(img, n,
                                                          img_ratio,
                                                          self.customstatus,
                                                          None,
                                                          None,
                                                          self.inputcolour6)
                    layout.name_text_place(img,
                                           firstnames[file_i],
                                           surnames[file_i], n,
                                           self.customstatus,
                                           self.inputtextsize,
                                           self.inputcolour5)
                    if self.inputlayout == 'Full-page grid layout':
                        layout.image_place(canvas_i, n, file_i, img,
                                           calc_size_output)
                    else:
                        layout.image_place(canvas_i, n, file_i, img, None)
                    prog_n += (100/n)/len(school)
                    self.signals.progress.emit(int(prog_n))

            # Add logo.
            if any([self.logo == [],
                    self.customstatus == True]):
                pass
            else:
                layout.logo_place(canvas_i, self.logo, n)

            # Add school and class name.
            if self.customstatus == False:
                layout.text_place(canvas_i,
                                  str(self.inputschool),
                                  str(grade[0][2]) + self.inputdate,
                                  logo_width, n, self.inputcolour4)

            # Save.
            if not self.customstatus:
                if any([self.inputsize == '18" x 12"',
                        self.inputsize == '12" x 8"']):
                    filename = grade[0][2] + '~PAN12'
                else:
                    filename = grade[0][2]
            else:
                filename = grade[0][2]
            for x in bad_characters:
                filename = filename.replace(x, ' ')
            savename = os.path.join(gridfolder, filename + '.jpg')
            canvas_i.save(savename,
                          quality = 95,
                          dpi = (300, 300),
                          icc_profile = img_profile)
            print(f'{filename}.jpg saved.')

            # Update metadata.
            img_iptc = IPTCInfo(savename, force = True)
            img_iptc['caption/abstract'] = 'PAN12'
            img_iptc.save()
            try:
                os.remove(savename + '~')
            except OSError:
                print('Error while deleting temporary files')

            # Save GoVisually version.
            if self.gvstatus == True:
                canvas_gv = layout.make_gv(canvas_i)
                canvas_gv.save(gvfolder+'\\'+ filename +'.jpg',
                               quality = 90,
                               dpi = (300, 300),
                               icc_profile = img_profile,
                               optimize = True,
                               subsampling = 2)

        if overflow_warning:
            overflow_warning_str = '\n'.join(overflow_warning)
            self.signals.error06.emit(len(overflow_warning),
                                      overflow_warning_str)
            raise UserWarning('Exception passed to user.')

        sys.exit()

    def excel_to_list(self, sheet, columns):
        wb = load_workbook(filename = sheet)
        ws = wb.active
        func_array = []
        for c in columns:
            func_list = []
            for cell in ws[c]:
                func_list.append(str(cell.value))
            func_array += [func_list]
        func_array = [list(a) for a in [*zip(*func_array)]]

        return func_array

    def csvwrite(self, path, rows, mode):
        with open(path, mode, newline = '\n') as csvfile:
            csvwriter = csv.writer(csvfile,
                                   delimiter = ',',
                                   quotechar = '"',
                                   quoting = csv.QUOTE_MINIMAL)
            csvwriter.writerow(rows)


class Window(QMainWindow):

    def __init__(self):
        super(Window, self).__init__()
        self.setWindowTitle('Grid Maker')
        self.setWindowIcon(QIcon(os.path.join(ROOT_DIR, 'gridmakericon.ico')))
        QApplication.setStyle(QStyleFactory.create('Fusion'))

        self.threadpool = QThreadPool()
        self.runner = JobRunner()
        self.runner.signals.loading.connect(self.update_loading)
        self.runner.signals.progress.connect(self.update_progress)
        self.runner.signals.error01.connect(self.input_error)
        self.runner.signals.error02.connect(self.file_error)
        self.runner.signals.error03.connect(self.spreadsheet_error)
        self.runner.signals.error04.connect(self.ratio_error)
        self.runner.signals.error05.connect(self.spreadsheet_duplicate_error)
        self.runner.signals.error06.connect(self.overflow_error)
        self.runner.signals.error07.connect(self.sitting_type_error)
        self.runner.signals.error08.connect(self.name_duplicate_error)
        self.runner.signals.go.connect(self.grid_choice)
        
        self.load = QProgressDialog('Checking files.', 'Cancel', 0, 100)
        self.load.setWindowTitle('Grid Maker')
        self.load.setWindowIcon(QIcon(os.path.join(
            ROOT_DIR, 'gridmakericon.ico')))
        self.load.setMinimumDuration(0)
        self.load.setCancelButton(None)

        self.selected_classes = set()
        
        self.threadpool.start(self.runner)

    def update_loading(self, n):
        self.load.setValue(n)
        QApplication.processEvents()

    def input_error(self):
        ie_w = QMessageBox.warning(self, 'Warning',
                                   'Invalid input.\n'+
                                   '\nPlease drag a folder onto the app.',
                                   QMessageBox.Ok)
                    
        if ie_w == QMessageBox.Ok:
            sys.exit()
        else:
            pass
        
    def file_error(self):
        fe_w = QMessageBox.warning(self, 'Warning',
                                   'Something is missing from the folder.\n'+
                                   '\nPlease check that the folder contains '+
                                   'both\nthe headshot images and the '+
                                   'sitting spreadsheet.',
                                   QMessageBox.Ok)
                    
        if fe_w == QMessageBox.Ok:
            sys.exit()
        else:
            pass

    def spreadsheet_error(self):
        se_w = QMessageBox.warning(self, 'Warning',
                                   'The spreadsheet is missing data for some '+
                                   'images.\nPlease see the file entitled '+
                                   '"Missing data.csv",\nupdate the '+
                                   'spreadsheet and try again.',
                                   QMessageBox.Ok)

        if se_w == QMessageBox.Ok:
            sys.exit()
        else:
            pass

    def spreadsheet_duplicate_error(self):
        sde_w = QMessageBox.warning(self, 'Warning',
                                   'The spreadsheet has duplicate student '+
                                   'IDs.\nPlease remove duplicate lines '+
                                   'and try again.',
                                   QMessageBox.Ok)

        if sde_w == QMessageBox.Ok:
            sys.exit()
        else:
            pass

    def ratio_error(self):
        se_w = QMessageBox.warning(self, 'Warning',
                                   'There is a problem with the image aspect '+
                                   'ratios.\n\nPlease either crop all images '+
                                   'to 3:2 ratio, or all images to 4:3 ratio.',
                                   QMessageBox.Ok)

        if se_w == QMessageBox.Ok:
            sys.exit()
        else:
            pass

    def overflow_error(self, n, i):
        if n == 1:
            n_text = 'grid was'
        else:
            n_text = 'grids were'
        o_w = QMessageBox.warning(self, 'Warning',
                                  f'Please note, the following {n_text} not '+
                                  'made\ndue to there being over 40 images:'+
                                  f'\n\n{i}',
                                  QMessageBox.Ok)

        if o_w == QMessageBox.Ok:
            sys.exit()
        else:
            pass

    def sitting_type_error(self):
        st_w = QMessageBox.warning(self, 'Warning',
                                   'There are unrecognised sitting types on '+
                                   'the spreadsheet.\n\nPlease correct the '+
                                   'sitting types so that staff can be '+
                                   'detected.',
                                   QMessageBox.Ok)

        if st_w == QMessageBox.Ok:
            sys.exit()
        else:
            pass

    def name_duplicate_error(self, n, i):
        if n == 1:
            n_text = 'individual appears'
        else:
            n_text = 'individuals appear'
        n_w = QMessageBox.warning(self, 'Warning',
                                  f'Please note, the following {n_text} '+
                                  'more than once in their class. \nPlease '+
                                  f'check for duplicates:\n\n{i}',
                                  QMessageBox.Ok)

        if n_w == QMessageBox.Ok:
            sys.exit()
        else:
            pass

    def grid_choice(self):
        self.setGeometry(670, 300, 340, 450)
        self.gridchoiceLabel = QLabel('Please select the grids you would like '
                                      + 'to make,\nthen choose a layout style.',
                                      self)
        self.gridchoiceLabel.move(0, 20)
        self.gridchoiceLabel.resize(340, 35)
        self.gridchoiceLabel.setAlignment(Qt.AlignCenter)
        
        self.listWidget = QListWidget(self)
        self.listWidget.setGeometry(40, 80, 260, 280)
        self.listWidget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.listWidget.setSortingEnabled(True)
        self.listWidget.addItems(list(self.runner.class_set))
        self.listWidget.selectAll()

        self.preformatBtn = QPushButton('Preformatted', self)
        self.preformatBtn.move(40, 390)
        self.preformatBtn.resize(120, 30)
        self.preformatBtn.clicked.connect(self.preformatted_grid)

        self.customBtn = QPushButton('Custom size', self)
        self.customBtn.clicked.connect(self.custom_grid)
        self.customBtn.move(180, 390)
        self.customBtn.resize(120, 30)

        self.show()

    def preformatted_grid(self):
        for i in self.listWidget.selectedItems():
            self.selected_classes.add(i.text())
        self.close()
        self.gridchoiceLabel.setParent(None)
        self.listWidget.setParent(None)
        self.preformatBtn.setParent(None)
        self.customBtn.setParent(None)
        
        self.setGeometry(600, 200, 480, 680)
        self.layoutChoice1 = QLabel('Size:', self)
        self.layoutChoice1.move(80, 60)
        self.layoutChoice2 = QLabel('Layout style:', self)
        self.layoutChoice2.move(80, 100)
        self.colourChoice3 = QLabel('Headshot order:', self)
        self.colourChoice3.move(80, 140)
        self.nameChoice = QLabel('Add names:', self)
        self.nameChoice.move(80, 180)
        self.schoolName = QLabel('School name:', self)
        self.schoolName.move(80, 220)
        self.dateFormat = QLabel('Date format:', self)
        self.dateFormat.move(80, 260)
        
        self.colourChoice1 = QLabel('Caption bar colour:', self)
        self.colourChoice1.move(80, 340)
        self.colourChoice2 = QLabel('Inner border colour:', self)
        self.colourChoice2.move(80, 380)
        self.colourChoice3 = QLabel('Outer border colour:', self)
        self.colourChoice3.move(80, 420)
        self.colourChoice4 = QLabel('Caption text colour:', self)
        self.colourChoice4.move(268, 340)
        self.colourChoice5 = QLabel('Name text colour:', self)
        self.colourChoice5.move(268, 380)
        self.colourChoice5.setStyleSheet('color: #808080;')
        self.colourChoice6 = QLabel('Keyline colour:', self)
        self.colourChoice6.move(268, 420)

        self.gvChoice = QLabel('with GoVisually versions:', self)
        self.gvChoice.move(217, 539)
        self.gvChoice.adjustSize()

        self.comboBox1 = QComboBox(self)
        self.comboBox1.move(250, 65)
        self.comboBox1.resize(150, 22)
        self.comboBox1.addItem('18" x 12"')
        self.comboBox1.addItem('12" x 8"')
        self.comboBox1.addItem('15" x 12"')
        self.comboBox1.addItem('10" x 8"')
        self.comboBox1.addItem('30" x 20"')
        
        self.comboBox2 = QComboBox(self)
        self.comboBox2.move(250, 105)
        self.comboBox2.resize(150, 22)
        self.comboBox2.addItem('Centred four-row layout')
        self.comboBox2.addItem('Full-page grid layout')

        self.comboBox3 = QComboBox(self)
        self.comboBox3.move(250, 145)
        self.comboBox3.resize(150, 22)
        self.comboBox3.addItem('First name')
        self.comboBox3.addItem('Surname')
        self.comboBox3.addItem('Student ID')
        self.comboBox3.addItem('Custom (column AT)')
        self.comboBox3.setCurrentText('Custom (column AT)')

        self.checkbox1 = QCheckBox(self)
        self.checkbox1.move(250, 182)
        self.checkbox1.setCheckState(Qt.Unchecked)
        self.checkbox1.toggled.connect(self.blank_colour)

        self.textbox1 = QLineEdit(self)
        self.textbox1.move(250, 225)
        self.textbox1.resize(150, 22)

        self.textbox2 = QLineEdit(self)
        self.textbox2.move(250, 265)
        self.textbox2.resize(150, 22)
        self.textbox2.setText(
            open(os.path.join(ROOT_DIR, 'year_text.txt')).read())

        self.btn1 = QPushButton(self)
        self.btn1.move(193, 345)
        self.btn1.resize(22, 22)
        self.btn1.clicked.connect(lambda: self.colour_picker(1))

        self.btn2 = QPushButton(self)
        self.btn2.move(193, 385)
        self.btn2.resize(22, 22)
        self.btn2.clicked.connect(lambda: self.colour_picker(2))

        self.btn3 = QPushButton(self)
        self.btn3.move(193, 425)
        self.btn3.resize(22, 22)
        self.btn3.clicked.connect(lambda: self.colour_picker(3))

        self.btn4 = QPushButton(self)
        self.btn4.move(378, 345)
        self.btn4.resize(22, 22)
        self.btn4.clicked.connect(lambda: self.colour_picker(4))

        self.btn5 = QPushButton(self)
        self.btn5.move(378, 385)
        self.btn5.resize(22, 22)
        self.btn5.clicked.connect(lambda: self.colour_picker(5))
        self.btn5.setEnabled(False)

        self.btn6 = QPushButton(self)
        self.btn6.move(378, 425)
        self.btn6.resize(22, 22)
        self.btn6.clicked.connect(lambda: self.colour_picker(6))

        self.checkbox2 = QCheckBox(self)
        self.checkbox2.move(346, 531)
        self.checkbox2.setCheckState(Qt.Unchecked)

        self.btn7 = QPushButton('Make grids', self)
        self.btn7.move(120, 530)
        self.btn7.resize(80, 30)
        self.btn7.clicked.connect(self.colour_check)

        self.progress = QProgressBar(self)
        self.progress.setGeometry(120, 590, 240, 18)

        if self.runner.logostatus == True:
            self.logoIndicator = QLabel('School logo detected', self)
            self.logoIndicator.move(20, 653)
            self.logoIndicator.adjustSize()
        else:
            self.logoIndicator = QLabel('No school logo detected', self)
            self.logoIndicator.move(20, 653)
            self.logoIndicator.adjustSize()

        self.show()

    def custom_grid(self):
        for i in self.listWidget.selectedItems():
            self.selected_classes.add(i.text())
        self.close()
        self.gridchoiceLabel.setParent(None)
        self.listWidget.setParent(None)
        self.preformatBtn.setParent(None)
        self.customBtn.setParent(None)
        
        self.setGeometry(600, 200, 480, 680)
        self.layoutChoice = QLabel('Headshot order:', self)
        self.layoutChoice.move(80, 60)
        self.layoutChoice1 = QLabel('Add names:', self)
        self.layoutChoice1.move(80, 110)
        self.layoutChoice2 = QLabel('Name text size:', self)
        self.layoutChoice2.move(80, 160)
        self.layoutChoice2.setStyleSheet('color: #808080;')
        self.sizeChoice1 = QLabel('Canvas width:', self)
        self.sizeChoice1.move(80, 210)
        self.sizeChoice2 = QLabel('Canvas height:', self)
        self.sizeChoice2.move(80, 260)
        self.layoutChoice3 = QLabel('pt', self)
        self.layoutChoice3.move(360, 160)
        self.layoutChoice3.setStyleSheet('color: #808080;')
        self.sizeChoice2 = QLabel('pixels', self)
        self.sizeChoice2.move(360, 210)
        self.sizeChoice2 = QLabel('pixels', self)
        self.sizeChoice2.move(360, 260)
        self.sizeChoice3 = QLabel('Image width padding:', self)
        self.sizeChoice3.move(80, 319)
        self.sizeChoice3.adjustSize() 
        self.sizeChoice4 = QLabel('Image height padding:', self)
        self.sizeChoice4.move(80, 369)
        self.sizeChoice4.adjustSize()
        self.sizeChoice4 = QLabel('x width', self)
        self.sizeChoice4.move(360, 319)
        self.sizeChoice4.adjustSize()
        self.sizeChoice4 = QLabel('x width', self)
        self.sizeChoice4.move(360, 369)
        self.sizeChoice4.adjustSize()
        self.colourChoice6 = QLabel('Keyline colour:', self)
        self.colourChoice6.move(80, 450)
        self.colourChoice6.adjustSize()
        self.colourChoice5 = QLabel('Name text colour:', self)
        self.colourChoice5.move(268, 450)
        self.colourChoice5.adjustSize()
        self.colourChoice5.setStyleSheet('color: #808080;')

        self.comboBox1 = QComboBox(self)
        self.comboBox1.move(250, 65)
        self.comboBox1.resize(150, 22)
        self.comboBox1.addItem('First name')
        self.comboBox1.addItem('Surname')
        self.comboBox1.addItem('Student ID')
        self.comboBox1.addItem('Custom (column AT)')
        self.comboBox1.setCurrentText('Surname')

        self.checkbox1 = QCheckBox(self)
        self.checkbox1.move(250, 112)
        self.checkbox1.setCheckState(Qt.Unchecked)
        self.checkbox1.toggled.connect(self.blank_colour_text)

        self.textbox0 = QLineEdit(self)
        self.textbox0.move(250, 165)
        self.textbox0.resize(100, 22)
        self.textbox0.setEnabled(False)

        self.textbox1 = QLineEdit(self)
        self.textbox1.move(250, 215)
        self.textbox1.resize(100, 22)

        self.textbox2 = QLineEdit(self)
        self.textbox2.move(250, 265)
        self.textbox2.resize(100, 22)

        self.textbox3 = QLineEdit(self)
        self.textbox3.move(250, 315)
        self.textbox3.resize(100, 22)
        self.textbox3.setText('0.2')

        self.textbox4 = QLineEdit(self)
        self.textbox4.move(250, 365)
        self.textbox4.resize(100, 22)
        self.textbox4.setText('0.2')

        self.btn6 = QPushButton(self)
        self.btn6.move(193, 445)
        self.btn6.resize(22, 22)
        self.btn6.clicked.connect(lambda: self.colour_picker(6))

        self.btn5 = QPushButton(self)
        self.btn5.move(378, 445)
        self.btn5.resize(22, 22)
        self.btn5.clicked.connect(lambda: self.colour_picker(5))
        self.btn5.setEnabled(False)

        self.btn1 = QPushButton('Make grids', self)
        self.btn1.move(198, 530)
        self.btn1.resize(80,30)
        self.btn1.clicked.connect(self.custom_grid_final)

        self.progress = QProgressBar(self)
        self.progress.setGeometry(130, 590, 220, 18)

        self.show()

    def blank_colour(self):
        if self.checkbox1.isChecked():
            self.btn5.setEnabled(True)
            self.colourChoice5.setStyleSheet('color: #000000;')
        else:
            self.btn5.setEnabled(False)
            self.colourChoice5.setStyleSheet('color: #808080;')

    def blank_colour_text(self):
        if self.checkbox1.isChecked():
            self.textbox0.setEnabled(True)
            self.btn5.setEnabled(True)
            self.layoutChoice2.setStyleSheet('color: #000000;')
            self.layoutChoice3.setStyleSheet('color: #000000;')
            self.colourChoice5.setStyleSheet('color: #000000;')
        else:
            self.textbox0.setEnabled(False)
            self.btn5.setEnabled(False)
            self.layoutChoice2.setStyleSheet('color: #808080;')
            self.layoutChoice3.setStyleSheet('color: #808080;')
            self.colourChoice5.setStyleSheet('color: #808080;')

    def colour_picker(self, col_n):
        colour = QColorDialog.getColor()
        button_string = (f'self.btn{col_n}.setStyleSheet('
                         + '"QWidget{background-color: %s}"'
                         + ' % colour.name())')
        output_string = f'self.runner.inputcolour{col_n} = colour.name()'
        eval(button_string)
        exec(output_string)

    def colour_error(self):
        cc_w = QMessageBox.warning(self, 'Warning',
                                   'Colours not selected.',
                                   QMessageBox.Ok)

        if cc_w == QMessageBox.Ok:
            pass
        else:
            pass

    def custom_grid_error(self):
        cg_w = QMessageBox.warning(self, 'Warning',
                                   'Please complete the form.',
                                   QMessageBox.Ok)

        if cg_w == QMessageBox.Ok:
            pass
        else:
            pass

    def colour_check(self):
        if any([self.runner.inputcolour1 == [],
                self.runner.inputcolour2 == [],
                self.runner.inputcolour3 == [],
                self.runner.inputcolour4 == [],
                self.runner.inputcolour6 == []]):
            self.colour_error()

        elif all([self.checkbox1.isChecked(),
                  self.runner.inputcolour5 == []]):
            self.colour_error()
            
        else:
            self.btn1.setEnabled(False)
            self.btn2.setEnabled(False)
            self.btn3.setEnabled(False)
            self.btn4.setEnabled(False)
            self.btn5.setEnabled(False)
            self.btn6.setEnabled(False)
            self.btn7.setEnabled(False)
            self.runner.inputsize = self.comboBox1.currentText()
            self.runner.inputlayout = self.comboBox2.currentText()
            self.runner.inputorder = self.comboBox3.currentText()
            self.runner.inputname = self.checkbox1.isChecked()
            self.runner.inputschool = self.textbox1.text()
            self.runner.inputdate = self.textbox2.text()
            self.runner.gvstatus = self.checkbox2.isChecked()
            self.runner.class_set = self.selected_classes
            
            with open(os.path.join(ROOT_DIR, 'year_text.txt'), 'w') as f:
                f.write(self.textbox2.text())
                
            self.runner.commence()

    def custom_grid_final(self):
        if any([self.textbox1.text() == '',
                self.textbox2.text() == '',
                self.textbox3.text() == '',
                self.textbox4.text() == '']):
            self.custom_grid_error()
        elif self.runner.inputcolour6 == []:
            self.colour_error()
        elif all([self.checkbox1.isChecked(),
                  self.runner.inputcolour5 == []]):
            self.colour_error()
        elif all([self.checkbox1.isChecked(),
                  self.textbox0.text() == '']):
            self.custom_grid_error()
            
        else:
            self.btn1.setEnabled(False)
            self.btn5.setEnabled(False)
            self.btn6.setEnabled(False)
            self.runner.inputtextsize = self.textbox0.text()
            self.runner.inputwidth = self.textbox1.text()
            self.runner.inputheight = self.textbox2.text()
            self.runner.inputwidthpad = self.textbox3.text()
            self.runner.inputheightpad = self.textbox4.text()
            self.runner.inputorder = self.comboBox1.currentText()
            self.runner.inputname = self.checkbox1.isChecked()
            self.runner.inputcolour1 = None
            self.runner.inputcolour2 = None
            self.runner.inputcolour3 = None
            self.runner.inputsize = '18" x 12"'
            self.runner.inputlayout = 'Full-page grid layout'
            self.runner.customstatus = True
            self.runner.class_set = self.selected_classes
            
            self.runner.commence()

    def update_progress(self, n):
        self.progress.setValue(n)
        QApplication.processEvents()


if __name__ == "__main__":
    ROOT_DIR = os.path.realpath(os.path.dirname(__file__))
    def run():
        sys.excepthook = except_hook  # for showing exception warnings
        app = QApplication(sys.argv)
        Gui = Window()
        sys.exit(app.exec_())
        print('End of program')
    def except_hook(cls, exception, traceback):  #
        sys.__excepthook__(cls, exception, traceback) #

run()


